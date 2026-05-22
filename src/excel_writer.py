"""Excel生成：将提取的表格写入Excel文件，智能命名"""

import re
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from src.table_extractor import ExtractedTable


HEADER_FILL = PatternFill(start_color="B8CCE4", end_color="B8CCE4", fill_type="solid")
HEADER_FONT = Font(bold=True, size=11)
TITLE_FONT = Font(bold=True, size=14)
NORMAL_FONT = Font(size=11)
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
ALT_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")


def write_tables_to_excel(
    tables: list[ExtractedTable],
    output_dir: Path,
    company_name: str,
    filing_title: str,
    report_year: str,
    doc_type: str,
) -> list[Path]:
    """
    将所有表格写入单个Excel文件，每个表格一个sheet。

    Returns:
        生成的Excel文件路径列表（只有一个文件）
    """
    output_dir.mkdir(parents=True, exist_ok=True)
    path = _generate_filepath(output_dir, company_name, report_year, doc_type, "全部表格")
    _write_workbook(path, tables, None)
    return [path]


def _generate_filepath(
    output_dir: Path,
    company_name: str,
    year: str,
    doc_type: str,
    title: str | None,
) -> Path:
    """生成智能文件名"""
    safe_title = _sanitize(title or "表格")[:60]
    safe_doc_type = _sanitize(doc_type)[:40]
    base = f"{company_name}_{year}_{safe_doc_type}_{safe_title}.xlsx"
    path = output_dir / base

    # 防止重名
    counter = 1
    while path.exists():
        path = output_dir / f"{company_name}_{year}_{doc_type}_{safe_title}_{counter}.xlsx"
        counter += 1
    return path


def _write_workbook(
    path: Path,
    tables: list[ExtractedTable],
    workbook_title: str | None,
):
    """写入一个Excel工作簿，每个表格一个sheet"""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    used_names: set[str] = set()

    for i, table in enumerate(tables):
        sheet_title = _sheet_name(table, i)
        # 防止重名
        original = sheet_title
        counter = 2
        while sheet_title in used_names:
            suffix = f"_{counter}"
            sheet_title = (original[:31 - len(suffix)] + suffix)
            counter += 1
        used_names.add(sheet_title)
        ws = wb.create_sheet(title=sheet_title)
        _write_sheet(ws, table.data, table.title)

    wb.save(path)


def _sheet_name(table: ExtractedTable, idx: int) -> str:
    """生成sheet名称（不超过31字符）"""
    if table.title:
        name = _sanitize_sheet(table.title)[:31]
    else:
        name = f"第{table.page_number}页_表{table.table_index_on_page + 1}"
    return name


def _write_sheet(ws, data: list[list[str]], title: str | None):
    """格式化写入一个sheet"""
    start_row = 1

    # 标题行
    if title:
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(1, len(data[0]) if data else 1))
        cell = ws.cell(row=1, column=1, value=title)
        cell.font = TITLE_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        start_row = 2

    # 表头行
    if data:
        for col_idx, val in enumerate(data[0], 1):
            cell = ws.cell(row=start_row, column=col_idx, value=val)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = THIN_BORDER

    # 数据行
    for row_idx, row in enumerate(data[1:], start_row + 1):
        is_alt = (row_idx - start_row) % 2 == 0
        for col_idx, val in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = NORMAL_FONT
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            if is_alt:
                cell.fill = ALT_FILL

    # 自动列宽
    _auto_column_width(ws, data)

    # 冻结首行（标题+表头）
    ws.freeze_panes = ws.cell(row=start_row + 1, column=1)


def _auto_column_width(ws, data: list[list[str]]):
    """自动调整列宽"""
    if not data:
        return
    max_cols = max(len(r) for r in data) if data else 0
    for col_idx in range(1, max_cols + 1):
        max_width = 0
        for row in data:
            if col_idx <= len(row):
                # 估算中文字符宽度
                text = str(row[col_idx - 1])
                width = sum(2 if ord(c) > 127 else 1 for c in text)
                max_width = max(max_width, width)
        # 限制宽度
        col_width = min(max_width + 4, 60)
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width


def _sanitize(name: str) -> str:
    """清理文件名中非法字符"""
    name = re.sub(r'[<>:"/\\|?*\n\r\t]', "", name)
    return name.strip().strip(".")


def _sanitize_sheet(name: str) -> str:
    r"""清理sheet名中的非法字符（Excel sheet名不能含: \ / * ? : [ ]）"""
    name = re.sub(r'[\\/*?:\[\]\n\r\t]', "", name)
    return name.strip().strip(".")
